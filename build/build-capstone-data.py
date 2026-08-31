#!/usr/bin/env python3
"""Build the seven Day 3 capstone scenario workbooks — one self-contained file per team.

site/alshaya-capstone-s1.xlsx … s7.xlsx. Each workbook holds every sheet its
scenario needs (including copies of the relevant Day 2 source sheets), so a team
uploads ONE file. Each workbook carries exactly one planted inconsistency,
catchable inside the file itself, documented in
instructor/coded-alshaya-capstone-data-key.html.

Needs openpyxl. Reads site/alshaya-day-2-dataset.xlsx as the source of truth.
"""
import random
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
SITE = ROOT / 'site'
SRC = openpyxl.load_workbook(SITE / 'alshaya-day-2-dataset.xlsx')

TITLE = Font(bold=True, size=13)
NOTE = Font(italic=True, size=10)
HEAD = Font(bold=True)

MONTHS = ['2025-07', '2025-08', '2025-09', '2025-10', '2025-11', '2025-12',
          '2026-01', '2026-02', '2026-03', '2026-04', '2026-05', '2026-06']


def rows_of(name):
    return [r for r in SRC[name].iter_rows(min_row=5, values_only=True) if r[0]]


def sheet(wb, name, title, note, widths):
    ws = wb.create_sheet(name)
    ws['A1'] = title; ws['A1'].font = TITLE
    ws['A2'] = note; ws['A2'].font = NOTE
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    return ws


def header(ws, row, cols):
    for i, c in enumerate(cols, 1):
        ws.cell(row=row, column=i, value=c).font = HEAD


def put(ws, row, values):
    for j, v in enumerate(values, 1):
        ws.cell(row=row, column=j, value=v)


def copy_source(wb, src_name, new_name=None):
    src = SRC[src_name]
    ws = wb.create_sheet(new_name or src_name)
    for col in 'ABCDEFGHIJ':
        if col in src.column_dimensions:
            ws.column_dimensions[col].width = src.column_dimensions[col].width
    for r, row in enumerate(src.iter_rows(values_only=True), 1):
        for c, v in enumerate(row, 1):
            if v is not None:
                cell = ws.cell(row=r, column=c, value=v)
                if r in (1, 2, 4):
                    cell.font = {1: TITLE, 2: NOTE, 4: HEAD}[r]
    return ws


def readme(wb, scode, sname, contents):
    ws = sheet(wb, 'README', f'{scode} · {sname} — capstone scenario workbook', '', [95])
    lines = [
        'SAMPLE DATA — TRAINING USE ONLY. Fictional portfolio; safe to upload to Claude.',
        'This is your team’s complete dataset for the Day 3 capstone. Everything you need is in',
        'this one file.',
        '',
        'Sheets in this workbook:',
    ] + [f'  • {c}' for c in contents] + [
        '',
        'ONE WARNING, WORTH READING TWICE:',
        'This workbook contains exactly ONE planted inconsistency — between its sheets, or',
        'against its own totals. Your verifier’s job is to catch it before it reaches your',
        'briefing. Flag it in your pitch; do not silently pick a side.',
    ]
    for i, line in enumerate(lines, 4):
        ws.cell(row=i, column=1, value=line)


# ---------------------------------------------------------------- source facts
tickets = rows_of('Maintenance Log')
budget = rows_of('Budget vs Actual')
survey = rows_of('Employee Survey')
cands = rows_of('Candidates')

ticket_ids = {t[0] for t in tickets}
june_closed = Counter(t[9] for t in tickets
                      if t[6] == 'Closed' and str(t[7]).startswith('2026-06'))
open_by_cat = defaultdict(list)
closed_apr_may = []
for t in tickets:
    if t[6] == 'Open':
        open_by_cat[t[3]].append(t)
    elif str(t[1]) < '2026-06':
        closed_apr_may.append(t)
budget_map = {r[0]: (r[1], r[2]) for r in budget}
print('June closures (truth):', dict(june_closed))


def new_wb():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    return wb


SLA_HRS = {'High': 4, 'Medium': 24, 'Low': 72}

# ================================================================ S1 · Facilities
wb = new_wb()
readme(wb, 'S1', 'The Contractor Turnaround', [
    'Maintenance Log — all 68 tickets, Apr–Jun 2026',
    'Response Times — first-response hours against SLA, per ticket',
    'SLA & Attendance — contracted service levels and the Apr–Jun site-attendance record'])
copy_source(wb, 'Maintenance Log')

ws = sheet(wb, 'Response Times', 'First-response times against SLA, per ticket',
           'Sample data — fictional. SLA targets: High 4h · Medium 24h · Low 72h. Join to the Maintenance Log by Ticket ID.',
           [12, 14, 12, 20, 20, 14])
header(ws, 4, ['Ticket ID', 'Priority', 'Contractor', 'SLA target (hours)', 'First response (hours)', 'SLA met'])
rngA = random.Random(11)
row = 5
for t in sorted(tickets, key=lambda x: str(x[1])):
    target = SLA_HRS[t[5]]
    who = t[9]
    late_bias = 0.15
    if who == '[CONTRACTOR B]':
        month = str(t[1])[5:7]
        late_bias = {'04': 0.25, '05': 0.55, '06': 0.9}.get(month, 0.25)
    late = rngA.random() < late_bias
    resp = round(target * rngA.uniform(1.2, 3.4)) if late else round(target * rngA.uniform(0.2, 0.95), 1)
    resp = max(resp, 1)
    put(ws, row, (t[0], t[5], who, target, resp, 'No' if resp > target else 'Yes'))
    row += 1

ws = sheet(wb, 'SLA & Attendance', 'Contractor SLA terms and Apr–Jun site attendance',
           'Sample data — fictional. Cross-check the June summary against the Maintenance Log sheet.',
           [26, 20, 20, 20, 34])
ws['A4'] = 'Service levels (contracted)'; ws['A4'].font = HEAD
header(ws, 5, ['Priority', 'Respond within', 'Close within', 'Penalty band', ''])
for i, r in enumerate([
        ('High', '4 hours', '3 days', '2% of monthly fee'),
        ('Medium', '1 working day', '5 days', '1% of monthly fee'),
        ('Low', '3 working days', '10 days', '0.5% of monthly fee')], 6):
    put(ws, i, r)
ws['A9'] = 'Penalties cap at 10% of the monthly fee. Site attendance: [CONTRACTOR B] Mon–Thu, [CONTRACTOR D] Sun–Wed.'
ws['A9'].font = NOTE
ws['A11'] = 'Site attendance by week (Apr–Jun 2026)'; ws['A11'].font = HEAD
header(ws, 12, ['Week commencing', 'Contractor', 'Days required', 'Days attended', 'Site log note'])
b_days = [4, 4, 3, 4,  4, 3, 3, 2,  3, 1, 0, 0, 0]   # 13 weeks; the fade-out
d_days = [4, 4, 4, 3,  4, 4, 4, 4,  4, 4, 3, 4, 4]
week_starts = ['2026-03-29', '2026-04-05', '2026-04-12', '2026-04-19', '2026-04-26',
               '2026-05-03', '2026-05-10', '2026-05-17', '2026-05-24', '2026-05-31',
               '2026-06-07', '2026-06-14', '2026-06-21']
b_notes = {8: 'Lift parts chase; left early twice', 9: 'Attended Mon only; no works signed off',
           10: 'No show; calls unanswered', 11: 'No show', 12: 'No show'}
row = 13
for i, wk in enumerate(week_starts):
    put(ws, row, (wk, '[CONTRACTOR B]', 4, b_days[i], b_notes.get(i, ''))); row += 1
    put(ws, row, (wk, '[CONTRACTOR D]', 4, d_days[i],
                  'One day lost to materials delay' if i == 10 else '')); row += 1
row += 1
ws.cell(row=row, column=1, value='June summary (as reported by the contract manager)').font = HEAD
header(ws, row + 1, ['Contractor', 'June days attended / required', 'Tickets closed in June'])
put(ws, row + 2, ('[CONTRACTOR B]', f'{sum(b_days[-4:])} / 16', 6))            # PLANTED (log: 1)
put(ws, row + 3, ('[CONTRACTOR D]', f'{sum(d_days[-4:])} / 16', june_closed.get('[CONTRACTOR D]', 0)))
put(ws, row + 4, ('In-house', 'n/a', june_closed.get('In-house', 0)))
wb.save(SITE / 'alshaya-capstone-s1.xlsx')
print('S1 planted: summary says B closed 6 in June; its own log shows', june_closed.get('[CONTRACTOR B]', 0))

# ================================================================ S2 · Finance
wb = new_wb()
readme(wb, 'S2', 'The Cost Control Programme', [
    'Budget vs Actual — the Q3 operating budget, planned against actual',
    'Cost Detail — twelve months of monthly planned and actual cost, by category',
    'Occupancy & Income — six months of occupancy, rent and collections'])
copy_source(wb, 'Budget vs Actual')

ws = sheet(wb, 'Cost Detail', 'Monthly operating cost by category — Jul 2025 to Jun 2026',
           'Sample data — fictional. The Apr–Jun months are the Q3 quarter in the Budget vs Actual sheet.',
           [12, 30, 16, 16, 40])
header(ws, 4, ['Month', 'Category', 'Planned (KWD)', 'Actual (KWD)', 'Note'])
rngB = random.Random(22)
# fixed Apr–Jun actuals (maintenance carries the planted +600 overstatement)
q3_actual = {
    'Property Maintenance': [14800, 16200, 21900],     # PLANTED: sums 52,900 vs 52,300
    'Marketing & Leasing':  [6100, 4900, 4400],
    'Utilities':            [7600, 7900, 8600],
    'Payroll & Contracted Staff': [31600, 31600, 31600],
    'General Admin':        [3220, 3220, 3210],
}
season = {'Utilities': {'2025-07': 1.25, '2025-08': 1.3, '2025-09': 1.15, '2026-05': 1.05, '2026-06': 1.15}}
notes = {('2026-06', 'Property Maintenance'): 'North lift parts + emergency call-outs',
         ('2026-05', 'General Admin'): 'Insurance renewal',
         ('2026-04', 'Marketing & Leasing'): 'Campaign paused pending upper-floor plan'}
row = 5
for m in MONTHS:
    for cat, (planned_q, actual_q) in budget_map.items():
        planned_m = round(planned_q / 3 / 10) * 10
        if m in MONTHS[-3:]:
            actual_m = q3_actual[cat][MONTHS[-3:].index(m)]
        else:
            base = planned_m * season.get(cat, {}).get(m, 1.0)
            actual_m = round(base * rngB.uniform(0.93, 1.09) / 10) * 10
        put(ws, row, (m, cat, planned_m, actual_m, notes.get((m, cat), '')))
        row += 1
ws.cell(row=row + 1, column=1,
        value='Monthly planned = quarterly planned ÷ 3, held flat across the year.').font = NOTE
copy_source(wb, 'Occupancy & Income')
wb.save(SITE / 'alshaya-capstone-s2.xlsx')
print('S2 planted: Apr+May+Jun maintenance actuals sum 52,900 vs Budget sheet quarter', budget_map['Property Maintenance'][1])

# ================================================================ S3 · Leasing
wb = new_wb()
readme(wb, 'S3', 'The Upper-Floor Push', [
    'Occupancy & Income — six months of occupancy, rent and collections',
    'Unit Vacancy Schedule — all 48 [PROPERTY A] units, with use and lease expiry',
    'Enquiries & Viewings — six months of leasing enquiries and their outcomes'])
copy_source(wb, 'Occupancy & Income')

ws = sheet(wb, 'Unit Vacancy Schedule', '[PROPERTY A] unit-by-unit vacancy schedule — quarter end',
           'Sample data — fictional. Cross-check the summary against the Occupancy & Income sheet (June).',
           [10, 12, 12, 12, 16, 16, 18, 14, 42])
ws['A4'] = 'Summary: 36 units let · 12 vacant · 75.0% occupancy · avg achieved rate KWD 11.40/sqm vs KWD 13.00 asking'
ws['A4'].font = HEAD
header(ws, 6, ['Unit', 'Floor', 'Size (sqm)', 'Status', 'Months vacant', 'Rate (KWD/sqm)',
               'Use', 'Lease expiry', 'Note'])
rng = random.Random(31)
uses = ['Fashion', 'Services', 'F&B', 'Electronics', 'Homeware', 'Pharmacy', 'Salon']
units = []
plan = [('G', 16, {8}), ('F1', 16, {3, 6, 9, 14}), ('F2', 16, {1, 4, 5, 7, 10, 12, 13, 16})]   # F1-09 PLANTED
long_vacant = {('F2', 1): 14, ('F2', 4): 11, ('F2', 7): 10, ('F2', 10): 12, ('F2', 13): 9}
for floor, count, vac in plan:
    for n in range(1, count + 1):
        vacant = n in vac
        size = 240 if (floor, n) == ('G', 12) else rng.choice([60, 75, 90, 110, 140, 180, 220, 260, 320])
        months = long_vacant.get((floor, n), rng.choice([1, 2, 3, 4, 6]) if vacant else 0)
        note_txt = 'New signing — F&B operator, 5-year term' if (floor, n) == ('G', 12) else ''
        use = '' if vacant else ('F&B' if (floor, n) == ('G', 12) else rng.choice(uses))
        expiry = '' if vacant else f'20{rng.choice([26, 27, 27, 28, 28, 29, 30])}-{rng.randint(1, 12):02d}'
        units.append([f'{floor}-{n:02d}', {'G': 'Ground', 'F1': 'First', 'F2': 'Upper'}[floor], size,
                      'Vacant' if vacant else 'Let', months if vacant else '', None, use, expiry, note_txt])
let_units = [u for u in units if u[3] == 'Let']
rngR = random.Random(310)
rates = [round(rngR.uniform(9.8, 12.9), 2) for _ in let_units]
diff = round(11.40 * len(let_units), 2) - round(sum(rates), 2)
step = round(diff / len(rates), 2)
rates = [round(r + step, 2) for r in rates]
rates[-1] = round(rates[-1] + (round(11.40 * len(let_units), 2) - round(sum(rates), 2)), 2)
assert 8 <= rates[-1] <= 14, rates[-1]
for u, r in zip(let_units, rates):
    u[5] = r
for u in units:
    if u[3] == 'Vacant':
        u[5] = 13.00
for i, u in enumerate(units, 7):
    put(ws, i, u)

ws = sheet(wb, 'Enquiries & Viewings', 'Leasing enquiries — Jan–Jun 2026',
           'Sample data — fictional. Every enquiry logged by the leasing team, with its outcome.',
           [12, 12, 14, 14, 16, 18, 30])
header(ws, 4, ['Ref', 'Month', 'Floor interest', 'Size wanted (sqm)', 'Source', 'Outcome', 'Note'])
rngE = random.Random(44)
monthly_enq = {'2026-01': 9, '2026-02': 8, '2026-03': 8, '2026-04': 7, '2026-05': 7, '2026-06': 9}
outcomes_g = ['Viewing held — negotiating', 'Viewing held — declined (rate)', 'Let (prior quarter)',
              'Viewing held — declined (fit-out cost)', 'No viewing — went elsewhere']
row = 5
ref = 1
for m, count in monthly_enq.items():
    for _ in range(count):
        fl = rngE.choices(['Ground', 'First', 'Upper'], weights=[6, 3, 1])[0]
        out = 'Let — F&B signing' if (m == '2026-05' and fl == 'Ground' and ref % 7 == 0) \
            else rngE.choice(outcomes_g)
        note_txt = 'Asked about footfall figures' if fl == 'Upper' else ''
        put(ws, row, (f'ENQ-{ref:03d}', m, fl, rngE.choice([60, 80, 100, 140, 200, 240, 300]),
                      rngE.choice(['Broker', 'Website', 'Walk-in', 'Referral']), out, note_txt))
        row += 1; ref += 1
wb.save(SITE / 'alshaya-capstone-s3.xlsx')
n_vac = sum(1 for u in units if u[3] == 'Vacant')
print(f'S3 planted: rows show {48-n_vac} let / {n_vac} vacant; summary and Occupancy sheet say 36/12; enquiries: {ref-1}')

# ================================================================ S4 · HR hiring
wb = new_wb()
readme(wb, 'S4', 'The Hiring Sprint', [
    'Role Requirements — agreed criteria for the four open roles',
    'Candidate Register — the HR system of record (40 candidates)',
    'Agency Long-List — the recruiting agency’s transcription of the same pool',
    'Applications by Week — eight weeks of application volume by source'])
ws = sheet(wb, 'Role Requirements', 'Role requirements (agreed with hiring managers)',
           'Sample data — fictional. [PROPERTY C] opens with 14 units; the facilities team must be staffed first.',
           [24, 16, 28, 20])
header(ws, 4, ['Role', 'Min experience', 'Must-have certification', 'Max notice (weeks)'])
for i, r in enumerate([
        ('Property Manager', '6 years', 'RICS (full or part)', 8),
        ('Leasing Executive', '3 years', '— (certification a plus)', 6),
        ('Facilities Technician', '5 years', 'HVAC Level 2+ or IOSH', 8),
        ('Facilities Manager', '8 years', 'NEBOSH or IOSH', 12)], 5):
    put(ws, i, r)

extra = [
    ('C-13', 'Property Manager', 7, 'Mixed-use portfolios, service charge budgets', 'RICS (part)', 6, 'Salmiya', 'Not screened'),
    ('C-14', 'Property Manager', 15, 'Regional asset management, refurbishments', 'RICS, PMP', 16, 'Kuwait City', 'Not screened'),
    ('C-15', 'Property Manager', 5, 'Residential blocks, tenant relations', '—', 4, 'Hawally', 'Not screened'),
    ('C-16', 'Leasing Executive', 8, 'Anchor-tenant negotiation, F&B leasing', 'Certified Leasing Prof.', 8, 'Kuwait City', 'Not screened'),
    ('C-17', 'Leasing Executive', 3, 'Retail leasing, CRM, reporting', '—', 2, 'Farwaniya', 'Not screened'),
    ('C-18', 'Leasing Executive', 1, 'Sales, viewings', '—', 2, 'Salmiya', 'Not screened'),
    ('C-19', 'Facilities Technician', 6, 'HVAC, BMS monitoring, permits to work', 'HVAC Level 2', 4, 'Jahra', 'Not screened'),
    ('C-20', 'Facilities Technician', 9, 'Lifts, pumps, contractor supervision', 'HVAC Level 3, IOSH', 10, 'Kuwait City', 'Not screened'),
    ('C-21', 'Facilities Technician', 4, 'Electrical, lighting, small works', 'IOSH', 3, 'Hawally', 'Not screened'),
    ('C-22', 'Facilities Manager', 10, 'Multi-site FM, energy programmes', 'NEBOSH', 8, 'Kuwait City', 'Not screened'),
    ('C-23', 'Facilities Manager', 8, 'Hard services, CAFM rollouts', 'IOSH, PMP', 6, 'Salmiya', 'Not screened'),
    ('C-24', 'Facilities Manager', 19, 'Airport & mall operations, capex', 'NEBOSH, RICS', 14, 'Kuwait City', 'Not screened'),
    ('C-25', 'Property Manager', 9, 'Retail centres, tenant mix strategy', 'RICS', 10, 'Kuwait City', 'Not screened'),
    ('C-26', 'Property Manager', 3, 'Assistant PM, service charges', '—', 3, 'Farwaniya', 'Not screened'),
    ('C-27', 'Property Manager', 11, 'Mixed-use, refurb delivery', 'RICS (part), IOSH', 7, 'Salmiya', 'Not screened'),
    ('C-28', 'Property Manager', 6, 'Office portfolios, lease admin', 'RICS (part)', 5, 'Hawally', 'Not screened'),
    ('C-29', 'Leasing Executive', 5, 'Retail leasing, incentives design', 'Certified Leasing Prof.', 5, 'Kuwait City', 'Not screened'),
    ('C-30', 'Leasing Executive', 4, 'F&B leasing, brand outreach', '—', 6, 'Salmiya', 'Not screened'),
    ('C-31', 'Leasing Executive', 2, 'Viewings, CRM hygiene', '—', 2, 'Jahra', 'Not screened'),
    ('C-32', 'Leasing Executive', 7, 'Anchor renewals, footfall analytics', 'Certified Leasing Prof.', 9, 'Kuwait City', 'Not screened'),
    ('C-33', 'Facilities Technician', 5, 'HVAC, chillers, BMS alarms', 'HVAC Level 2, IOSH', 6, 'Farwaniya', 'Not screened'),
    ('C-34', 'Facilities Technician', 12, 'Lifts specialist, OEM trained', 'HVAC Level 3', 9, 'Kuwait City', 'Not screened'),
    ('C-35', 'Facilities Technician', 3, 'General maintenance, plumbing', '—', 2, 'Hawally', 'Not screened'),
    ('C-36', 'Facilities Technician', 8, 'HVAC + electrical dual trade', 'HVAC Level 3, IOSH', 8, 'Salmiya', 'Not screened'),
    ('C-37', 'Facilities Manager', 13, 'Retail FM, contractor SLAs', 'NEBOSH, IOSH', 10, 'Kuwait City', 'Not screened'),
    ('C-38', 'Facilities Manager', 7, 'Soft services lead, moving to hard FM', 'IOSH', 4, 'Jahra', 'Not screened'),
    ('C-39', 'Facilities Manager', 9, 'Mixed-use towers, energy retrofits', 'NEBOSH', 12, 'Salmiya', 'Not screened'),
    ('C-40', 'Facilities Manager', 16, 'Hospital estates, statutory compliance', 'NEBOSH, PMP', 13, 'Kuwait City', 'Not screened'),
]
COLS = ['Candidate ID', 'Role Applied', 'Years Experience', 'Key Skills',
        'Certifications', 'Notice Period (weeks)', 'Location', 'Screening Status']
pool = list(cands) + extra
ws = sheet(wb, 'Candidate Register', 'Candidate register — HR system of record (40 candidates)',
           'Sample data — fictional candidates, IDs only. This sheet is the source of truth.',
           [12, 22, 14, 44, 26, 18, 16, 14])
header(ws, 4, COLS)
for i, c in enumerate(pool, 5):
    put(ws, i, list(c))
ws = sheet(wb, 'Agency Long-List', 'Agency long-list — as transcribed by the recruiting agency',
           'Sample data — fictional. The agency retyped the register; verify before screening from this sheet.',
           [12, 22, 14, 44, 26, 18, 16, 14])
header(ws, 4, COLS)
for i, c in enumerate(pool, 5):
    c = list(c)
    if c[0] == 'C-03':
        c[2] = 8                                    # PLANTED: register says 12
    put(ws, i, c)

ws = sheet(wb, 'Applications by Week', 'Applications received by week and source',
           'Sample data — fictional. The eight weeks to 30 June 2026; totals match the 40-candidate register.',
           [16, 12, 12, 12, 12])
header(ws, 4, ['Week commencing', 'Job boards', 'Agency', 'Referrals', 'Total'])
weekly = [(2, 1, 0), (3, 2, 0), (4, 2, 1), (3, 2, 0), (2, 3, 1), (3, 2, 1), (2, 2, 1), (2, 1, 0)]
assert sum(sum(w) for w in weekly) == 40
wk_starts = ['2026-05-04', '2026-05-11', '2026-05-18', '2026-05-25',
             '2026-06-01', '2026-06-08', '2026-06-15', '2026-06-22']
for i, (wk, w) in enumerate(zip(wk_starts, weekly), 5):
    put(ws, i, (wk, *w, sum(w)))
wb.save(SITE / 'alshaya-capstone-s4.xlsx')
print('S4 planted: agency long-list has C-03 at 8 years; the register says 12 · pool = 40')

# ================================================================ S5 · People
wb = new_wb()
readme(wb, 'S5', 'The Engagement Turnaround', [
    'Employee Survey — the 12 scored responses (2026)',
    'Survey Comments — the free-text comments collected with the survey',
    '2025 Survey — last year’s scores, for the trend',
    'Exit Interviews — leavers over the last twelve months, with stated reasons',
    'Training Hours — recorded training hours by department, last twelve months'])
copy_source(wb, 'Employee Survey')

ws = sheet(wb, 'Survey Comments', 'Employee survey — free-text comments',
           'Sample data — fictional. One comment per response. Cross-check IDs against the Employee Survey sheet.',
           [12, 14, 90])
header(ws, 4, ['Response ID', 'Department', 'Comment (verbatim)'])
comments = {
    'E01': 'Good team. I would like a clearer path to senior leasing roles.',
    'E02': 'Targets are fine but training budget requests keep getting deferred.',
    'E03': 'Best manager I have worked for. Keep the weekly one-to-ones.',
    'E04': 'We only hear about problems. Nobody has asked what we need to close tickets faster.',
    'E05': 'Shift handovers are chaotic. Growth here means leaving, honestly.',
    'E06': 'Proud of the work. The on-call rota needs another pair of hands.',
    'E07': 'Month-end is heavy but manageable. More systems training would help.',
    'E08': 'Good support from my manager. Cross-training with Operations would be welcome.',
    'E09': 'Reporting deadlines are fair. I would take a course if one were offered.',
    'E10': 'Too much firefighting. A development plan was promised last year.',
    'E11': 'Workload is fine. It is unclear how promotion decisions are made.',
    'E12': 'The team is friendly. I want more responsibility and have said so twice.',
}
dept_by_id = {r[0]: r[1] for r in survey}
row = 5
for rid, txt in comments.items():
    put(ws, row, (rid, dept_by_id[rid], txt)); row += 1
put(ws, row, ('E13', 'Leasing', 'Everything is excellent, no changes needed at all.'))   # PLANTED

ws = sheet(wb, '2025 Survey', 'Employee satisfaction survey — 2025 (prior year)',
           'Sample data — fictional. 12 responses, scored 1 (lowest) to 5 (highest). Anonymised IDs; not the same people as 2026.',
           [12, 14, 18, 16, 20, 18])
header(ws, 4, ['Response ID', 'Department', 'Overall Satisfaction', 'Manager Support', 'Growth Opportunities', 'Work-Life Balance'])
prior = [
    ('P01', 'Leasing', 4, 4, 4, 4), ('P02', 'Leasing', 4, 3, 4, 3), ('P03', 'Leasing', 5, 5, 4, 4),
    ('P04', 'Maintenance', 4, 3, 4, 4), ('P05', 'Maintenance', 3, 3, 3, 3), ('P06', 'Maintenance', 4, 4, 4, 3),
    ('P07', 'Finance', 4, 4, 4, 3), ('P08', 'Finance', 4, 4, 4, 4), ('P09', 'Finance', 4, 4, 3, 3),
    ('P10', 'HR', 4, 4, 4, 4), ('P11', 'HR', 3, 3, 4, 3), ('P12', 'HR', 4, 4, 4, 4),
]
for i, r in enumerate(prior, 5):
    put(ws, i, r)

ws = sheet(wb, 'Exit Interviews', 'Exit interviews — leavers, Jul 2025 to Jun 2026',
           'Sample data — fictional. Primary stated reason, as categorised by HR.',
           [12, 12, 16, 28, 44])
header(ws, 4, ['Ref', 'Month', 'Department', 'Primary reason', 'Interview note'])
for i, r in enumerate([
        ('X-01', '2025-08', 'Maintenance', 'No growth path', 'Offered technician-lead role elsewhere'),
        ('X-02', '2025-10', 'Leasing', 'Compensation', 'Counter-offer declined'),
        ('X-03', '2025-11', 'Maintenance', 'No growth path', 'Asked twice about training; none scheduled'),
        ('X-04', '2026-01', 'HR', 'Relocation', 'Family move'),
        ('X-05', '2026-02', 'Maintenance', 'Workload', 'On-call rota cited; exhausted'),
        ('X-06', '2026-03', 'Finance', 'No growth path', 'Wanted systems specialisation'),
        ('X-07', '2026-05', 'Maintenance', 'No growth path', 'Third Maintenance leaver this year citing growth'),
        ('X-08', '2026-06', 'Leasing', 'Compensation', 'Moved to a competitor')], 5):
    put(ws, i, r)

ws = sheet(wb, 'Training Hours', 'Recorded training hours by department — Jul 2025 to Jun 2026',
           'Sample data — fictional. Hours logged in the HR system, per department per month.',
           [12, 12, 14, 12, 10])
header(ws, 4, ['Month', 'Leasing', 'Maintenance', 'Finance', 'HR'])
rngT = random.Random(55)
for i, m in enumerate(MONTHS, 5):
    put(ws, i, (m, rngT.choice([4, 6, 8]), rngT.choice([0, 0, 2]), rngT.choice([4, 6]), rngT.choice([2, 4, 6])))
wb.save(SITE / 'alshaya-capstone-s5.xlsx')
print('S5 planted: phantom respondent E13 for a', len(survey), 'response survey')

# ================================================================ S6 · Comms
wb = new_wb()
readme(wb, 'S6', 'The Tenant Confidence Campaign', [
    'Complaint Log — six months of tenant complaints, Jan–Jun 2026 (60 in total)',
    'Maintenance Log — all 68 tickets, for verifying linked ticket references'])
ws = sheet(wb, 'Complaint Log', 'Tenant complaint log — January to June 2026',
           'Sample data — fictional. The ticketing system went live in April 2026; earlier complaints have no linked ticket. Linked tickets refer to the Maintenance Log sheet.',
           [12, 12, 15, 12, 12, 12, 13, 46])
header(ws, 4, ['Ref', 'Date', 'Property', 'Category', 'Channel', 'Escalated', 'Linked ticket', 'Summary'])
summaries = {
    'Lifts': 'North lift out again; deliveries rerouted',
    'AC / HVAC': 'Unit too warm; ticket raised, no update since',
    'Plumbing': 'Leak reported; follow-up requested',
    'Cleaning': 'Common-area cleaning standard slipping',
    'Noise': 'Out-of-hours works noise',
    'Parking': 'Contractor vehicles in tenant bays',
    'Other': 'General service concern',
}
rng2 = random.Random(6)
lift_open = [t[0] for t in open_by_cat.get('Lifts', [])][:4]
hvac_open = [t[0] for t in open_by_cat.get('HVAC', [])][:6]
plumb_open = [t[0] for t in open_by_cat.get('Plumbing', [])][:3]
phantom = 'MT-097'
assert phantom not in ticket_ids
june_cats = ([('Lifts', tid) for tid in lift_open] + [('AC / HVAC', tid) for tid in hvac_open]
             + [('Plumbing', tid) for tid in plumb_open]
             + [('Cleaning', ''), ('Noise', ''), ('Cleaning', ''), ('Parking', ''), ('Noise', ''),
                ('AC / HVAC', phantom),                                     # PLANTED
                ('Lifts', ''), ('Parking', ''), ('Cleaning', ''), ('Other', '')])
assert len(june_cats) == 23
early_link_pool = [t[0] for t in closed_apr_may]
monthly = {'2026-01': 6, '2026-02': 7, '2026-03': 7, '2026-04': 8, '2026-05': 9}
early_cat_names = ['AC / HVAC', 'Cleaning', 'Plumbing', 'Noise', 'Parking', 'Lifts', 'Other']
ref = 1
row = 5
phantom_ref = None
for m, count in monthly.items():
    days = sorted(rng2.sample(range(1, 28), count))
    for i in range(count):
        cat = rng2.choice(early_cat_names)
        tid = rng2.choice(early_link_pool) if m >= '2026-04' and rng2.random() < 0.5 else ''
        put(ws, row, (f'CMP-{ref:03d}', f'{m}-{days[i]:02d}',
                      '[PROPERTY B]' if ref % 3 else '[PROPERTY A]', cat,
                      rng2.choice(['Email', 'Phone', 'Portal']), 'No', tid, summaries[cat]))
        row += 1; ref += 1
june_days = sorted(rng2.sample(range(1, 29), 23))
escalated_june = {2, 9, 17}
for i, (cat, tid) in enumerate(june_cats):
    summary = summaries[cat] + ('; tenant asks about a rent concession' if i == 9 else '')
    if tid == phantom:
        phantom_ref = f'CMP-{ref:03d}'
    put(ws, row, (f'CMP-{ref:03d}', f'2026-06-{june_days[i]:02d}',
                  '[PROPERTY B]' if cat == 'Lifts' or i % 3 else '[PROPERTY A]', cat,
                  rng2.choice(['Email', 'Phone', 'Portal']),
                  'Yes — in writing' if i in escalated_june else 'No', tid, summary))
    row += 1; ref += 1
copy_source(wb, 'Maintenance Log')
wb.save(SITE / 'alshaya-capstone-s6.xlsx')
print(f'S6 planted: {phantom_ref} links to {phantom} — absent from its own Maintenance Log sheet · complaints: {ref-1}')

# ================================================================ S7 · Operations
wb = new_wb()
readme(wb, 'S7', 'The Launch Readiness Plan', [
    'Project Facts — the confirmed [PROPERTY C] fit-out position (from the quarterly report)',
    'Milestone Schedule — the contractor’s programme, rev 4',
    'Snag List — open and closed snags by unit and trade',
    'Fit-out Cost Tracker — committed cost against the contingency allowance',
    'Occupancy & Income — portfolio context, including the [PROPERTY C] rows'])
ws = sheet(wb, 'Project Facts', '[PROPERTY C] — confirmed project position',
           'Sample data — fictional. Extracted from §5 of the Quarterly Property Operations Report, dated 15 July 2026.',
           [40, 55])
header(ws, 4, ['Fact', 'Detail'])
for i, r in enumerate([
        ('Property', '[PROPERTY C] — 14-unit tower, fit-out phase'),
        ('Joinery delivered to site', '2 June 2026'),
        ('Fit-out duration', 'Contractor requires six weeks from joinery delivery to practical completion'),
        ('Cost impact', 'KWD 18,000 against a KWD 25,000 contingency allowance'),
        ('Contractor', '[CONTRACTOR B] — performance review outstanding'),
        ('Board position', 'Revised handover date proposed, not yet ratified'),
        ('Tenant position', 'Fit-out crew booked; paying storage costs in the interim')], 5):
    put(ws, i, r)
ws = sheet(wb, 'Milestone Schedule', '[PROPERTY C] fit-out milestone schedule (contractor issue, rev 4)',
           'Sample data — fictional. Cross-check actual dates against the Project Facts sheet.',
           [34, 16, 16, 20, 40])
header(ws, 4, ['Milestone', 'Planned', 'Actual / Forecast', 'Status', 'Note'])
for i, r in enumerate([
        ('Fit-out contract award', '2026-03-14', '2026-03-14', 'Complete', ''),
        ('Joinery order placed', '2026-04-07', '2026-04-07', 'Complete', '3-week quoted lead time'),
        ('Joinery delivered to site', '2026-04-28', '2026-05-26', 'Complete', 'Supplier delay'),   # PLANTED: facts say 2 June
        ('Fit-out works (6 weeks from delivery)', '—', 'in progress', 'In progress', 'Contractor requires six weeks from joinery delivery'),
        ('Practical completion', '2026-06-09', '2026-07-07', 'Forecast', 'Delivery date + six weeks'),
        ('Snagging & close-out', '—', '2026-07-13', 'Forecast', 'One week allowed'),
        ('Handover to tenant', '2026-06-16', '2026-07-14', 'Forecast', 'Subject to Board ratification'),
        ('Board ratification of revised date', '—', 'pending', 'Pending', 'Proposed, not yet ratified'),
        ('Tenant fit-out start', '—', 'pending', 'Pending', 'Tenant crew booked; storage costs running')], 5):
    put(ws, i, r)

ws = sheet(wb, 'Snag List', '[PROPERTY C] snag list — as inspected 30 June 2026',
           'Sample data — fictional. Raised during progress inspections; severity per the consultant.',
           [10, 10, 18, 14, 12, 44])
header(ws, 4, ['Snag', 'Unit', 'Trade', 'Severity', 'Status', 'Description'])
rngS = random.Random(77)
trades = ['Joinery', 'Electrical', 'HVAC', 'Plumbing', 'Finishes', 'Fire systems']
sev = ['Minor', 'Minor', 'Minor', 'Moderate', 'Moderate', 'Major']
descs = {'Joinery': 'Door alignment / trim gaps', 'Electrical': 'Sockets untested or unlabelled',
         'HVAC': 'Diffuser balancing outstanding', 'Plumbing': 'Pressure test pending',
         'Finishes': 'Paint / tiling touch-ups', 'Fire systems': 'Detector head not commissioned'}
for i in range(28):
    tr = rngS.choice(trades)
    put(ws, i + 5, (f'SN-{i+1:03d}', f'U-{rngS.randint(1, 14):02d}', tr, rngS.choice(sev),
                    'Closed' if rngS.random() < 0.4 else 'Open', descs[tr]))

ws = sheet(wb, 'Fit-out Cost Tracker', '[PROPERTY C] fit-out variation cost tracker',
           'Sample data — fictional. Committed variations against the KWD 25,000 contingency allowance.',
           [10, 34, 16, 16, 30])
header(ws, 4, ['Line', 'Variation', 'Committed (KWD)', 'Forecast (KWD)', 'Note'])
lines = [
    ('V-01', 'Joinery re-delivery and crane hire', 4200, 4200, 'Supplier delay recovery'),
    ('V-02', 'Extended site preliminaries (3 weeks)', 5400, 5400, 'Contractor time-related costs'),
    ('V-03', 'Out-of-hours working to recover programme', 3600, 4400, 'Forecast rises if slip continues'),
    ('V-04', 'Tenant storage cost contribution', 2800, 3600, 'Accruing weekly until handover'),
    ('V-05', 'Design change — unit U-07 partition', 2000, 2000, 'Approved'),
]
committed = sum(l[2] for l in lines)
assert committed == 18000, committed
for i, l in enumerate(lines, 5):
    put(ws, i, l)
put(ws, 11, ('', 'Total committed', committed, sum(l[3] for l in lines), 'Contingency allowance: KWD 25,000'))
ws.cell(row=11, column=2).font = HEAD
copy_source(wb, 'Occupancy & Income')
wb.save(SITE / 'alshaya-capstone-s7.xlsx')
print('S7 planted: schedule says joinery actual 26 May; its own Project Facts sheet says 2 June · committed 18,000 vs 25,000 ✓')

for n in range(1, 8):
    p = SITE / f'alshaya-capstone-s{n}.xlsx'
    print('written', p.name, p.stat().st_size, 'bytes')
