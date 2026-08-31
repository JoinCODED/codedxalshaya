# -*- coding: utf-8 -*-
"""Day 2 lab dataset — fictional portfolio, consistent with the Day 1 quarterly report.
   VALUES ONLY, no formulas: trainees compute the analysis themselves, and a formula
   written by openpyxl has no cached value, so Claude would read blanks on upload."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date, timedelta

OUT = "/Users/ayaalsaqaf/codedxalshaya/site/alshaya-day-2-dataset.xlsx"

HEAD_FILL = PatternFill("solid", fgColor="1C0E12")
HEAD_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFF")
BODY_FONT = Font(name="Arial", size=10)
TITLE_FONT = Font(name="Arial", size=13, bold=True)
NOTE_FONT = Font(name="Arial", size=9, italic=True, color="666666")
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(bottom=THIN)

wb = Workbook()

def sheet(name, headers, rows, widths, title=None, note=None, numfmt=None):
    ws = wb.create_sheet(name)
    r = 1
    if title:
        ws.cell(r, 1, title).font = TITLE_FONT; r += 1
    if note:
        ws.cell(r, 1, note).font = NOTE_FONT; r += 1
    if title or note:
        r += 1
    hr = r
    for c, h in enumerate(headers, 1):
        cell = ws.cell(hr, c, h)
        cell.font = HEAD_FONT; cell.fill = HEAD_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[hr].height = 20
    for i, row in enumerate(rows):
        for c, v in enumerate(row, 1):
            cell = ws.cell(hr + 1 + i, c, v)
            cell.font = BODY_FONT; cell.border = BORDER
            if numfmt and c in numfmt:
                cell.number_format = numfmt[c]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = ws.cell(hr + 1, 1)
    ws.auto_filter.ref = (f"A{hr}:{get_column_letter(len(headers))}{hr + len(rows)}")
    return ws

# ─────────────────────────────── 1. Occupancy & income
MONTHS = ["2026-01","2026-02","2026-03","2026-04","2026-05","2026-06"]
A_LET  = [40, 39, 38, 38, 37, 36]      # ends at 36/48 = 75.0%  (Day 1 report §3)
B_LET  = [57, 57, 56, 56, 56, 56]      # ends at 56/62 = 90.3%
A_RATE, B_RATE = 4491, 3884            # KWD per let unit per month
A_COLL = [98.1, 97.8, 97.4, 96.9, 96.5, 96.2]
B_COLL = [98.4, 98.2, 97.9, 97.6, 97.1, 96.8]
A_FOOT = [214000, 208500, 203000, 199500, 194000, 188500]

occ = []
for i, m in enumerate(MONTHS):
    occ.append([m, "[PROPERTY A]", "Retail centre", 48, A_LET[i],
                round(A_LET[i]/48*100, 1), A_LET[i]*A_RATE, A_COLL[i], A_FOOT[i]])
    occ.append([m, "[PROPERTY B]", "Mixed-use tower", 62, B_LET[i],
                round(B_LET[i]/62*100, 1), B_LET[i]*B_RATE, B_COLL[i], ""])
    occ.append([m, "[PROPERTY C]", "Under construction", 14, 0, 0.0, 0, "", ""])

sheet("Occupancy & Income",
      ["Month","Property","Type","Units","Units Let","Occupancy %","Monthly Rent (KWD)","Collection %","Footfall"],
      occ, [11,15,20,8,11,13,19,13,11],
      title="Occupancy and income by month",
      note="Sample data — fictional portfolio. Jan–Jun 2026. Footfall tracked at the retail asset only.",
      numfmt={7:"#,##0", 6:"0.0", 8:"0.0", 9:"#,##0"})

# ─────────────────────────────── 2. Maintenance log
# Closed-ticket day counts chosen so each category average and the blended average
# match the Day 1 report exactly (HVAC 14, Plumbing 9, Lifts 21, Lighting 6, Other 4; blended 12).
CLOSED = {
    "HVAC":     [9, 11, 12, 13, 15, 16, 17, 19],   # 8 tickets, mean 14
    "Plumbing": [4, 6, 8, 10, 12, 14],             # 6 tickets, mean 9
    "Lifts":    [12, 16, 20, 22, 26, 30],          # 6 tickets, mean 21
    "Lighting": [2, 4, 5, 6, 8, 11],               # 6 tickets, mean 6
    "Other":    [3, 5],                            # 2 tickets, mean 4
}
# Open at 30 Jun: 40 total, split per the report; 11 of them raised before 31 May (>30 days open).
OPEN_SPLIT = {"HVAC":17, "Plumbing":9, "Lifts":7, "Lighting":5, "Other":2}
OPEN_AGED  = {"HVAC":5,  "Plumbing":2, "Lifts":3, "Lighting":1, "Other":0}   # sums to 11

DESC = {
 "HVAC":["AC not cooling, unit 3F","Air handler noise","Thermostat unresponsive","Condenser leak",
         "Uneven cooling, east wing","Filter replacement overdue","Chiller alarm","Vent blocked"],
 "Plumbing":["Leak under sink","Blocked drain, ground floor","Low water pressure","WC cistern running",
             "Pipe corrosion, riser 2"],
 "Lifts":["North lift out of service","Door sensor fault","Lift call button dead","Levelling fault",
          "Emergency phone not connecting"],
 "Lighting":["Corridor lights out, 7F","Emergency light failed test","Car park lamp flickering",
             "Signage light out"],
 "Other":["Door closer broken","Signage panel loose","Ceiling tile stained"],
}
PRIO  = {"HVAC":"Medium","Plumbing":"Medium","Lifts":"High","Lighting":"Low","Other":"Low"}
CONTR = ["[CONTRACTOR B]","[CONTRACTOR D]","In-house"]

rows, n = [], 0
def tid():
    global n; n += 1; return f"MT-{n:03d}"

# closed
d0 = date(2026, 4, 1)
for cat, days in CLOSED.items():
    for k, dd in enumerate(days):
        raised = d0 + timedelta(days=(k*7 + len(cat)) % 60)
        closed = raised + timedelta(days=dd)
        prop = "[PROPERTY A]" if (k % 2 == 0) else "[PROPERTY B]"
        rows.append([tid(), raised.isoformat(), prop, cat, DESC[cat][k % len(DESC[cat])],
                     PRIO[cat], "Closed", closed.isoformat(), dd, CONTR[k % 3]])
# open
for cat, cnt in OPEN_SPLIT.items():
    aged = OPEN_AGED[cat]
    for k in range(cnt):
        raised = (date(2026,4,20) + timedelta(days=k*2)) if k < aged \
                 else (date(2026,6,3) + timedelta(days=(k*3) % 24))
        prop = "[PROPERTY A]" if (k % 3 == 0) else "[PROPERTY B]"
        rows.append([tid(), raised.isoformat(), prop, cat, DESC[cat][k % len(DESC[cat])],
                     PRIO[cat], "Open", "", (date(2026,6,30) - raised).days, CONTR[k % 3]])
rows.sort(key=lambda r: r[1])

sheet("Maintenance Log",
      ["Ticket ID","Date Raised","Property","Category","Description","Priority","Status",
       "Date Closed","Days Open","Contractor"],
      rows, [11,13,15,11,32,10,9,13,11,17],
      title="Maintenance ticket log",
      note="Sample data — fictional. Apr–Jun 2026. 'Days Open' for open tickets is measured to 30 June 2026.")

# ─────────────────────────────── 3. Budget vs actual
BUD = [
 ["Property Maintenance", 45000, 52300],
 ["Marketing & Leasing",  18000, 15400],
 ["Utilities",            22000, 24100],
 ["Payroll & Contracted Staff", 96000, 94800],
 ["General Admin",         9000,  9650],
]
sheet("Budget vs Actual",
      ["Category","Planned (KWD)","Actual (KWD)"],
      BUD, [30,16,16],
      title="Q3 property operating budget",
      note="Sample data — fictional. Variance is deliberately left for you to calculate.",
      numfmt={2:"#,##0", 3:"#,##0"})

# ─────────────────────────────── 4. Employee survey
SURVEY = [
 ["E01","Leasing",4,4,3,4],["E02","Leasing",3,3,2,3],["E03","Leasing",5,5,4,4],
 ["E04","Maintenance",2,2,2,3],["E05","Maintenance",3,3,3,3],["E06","Maintenance",4,4,3,4],
 ["E07","Finance",4,5,3,3],["E08","Finance",3,3,4,2],["E09","Finance",5,4,5,4],
 ["E10","HR",4,4,4,5],["E11","HR",3,3,3,4],["E12","HR",2,3,2,2],
]
sheet("Employee Survey",
      ["Response ID","Department","Overall Satisfaction","Manager Support",
       "Growth Opportunities","Work-Life Balance"],
      SURVEY, [13,15,20,17,21,19],
      title="Employee satisfaction survey",
      note="Sample data — fictional. 12 responses, scored 1 (lowest) to 5 (highest).")

# ─────────────────────────────── 5. Candidates (HR screening)
CAND = [
 ["C-01","Property Manager",9,"Leasing, tenant relations, budgeting","RICS (part)",8,"Kuwait City","Not screened"],
 ["C-02","Property Manager",4,"Tenant relations, reporting","—",4,"Hawally","Not screened"],
 ["C-03","Property Manager",12,"Asset management, leasing, team lead","RICS, MBA",12,"Kuwait City","Not screened"],
 ["C-04","Leasing Executive",2,"Sales, CRM, viewings","—",2,"Salmiya","Not screened"],
 ["C-05","Leasing Executive",6,"Retail leasing, negotiation, CRM","Certified Leasing Prof.",6,"Kuwait City","Not screened"],
 ["C-06","Leasing Executive",1,"Customer service, admin","—",2,"Farwaniya","Not screened"],
 ["C-07","Facilities Technician",7,"HVAC, electrical, preventive maintenance","HVAC Level 3",4,"Jahra","Not screened"],
 ["C-08","Facilities Technician",3,"Plumbing, general maintenance","—",2,"Hawally","Not screened"],
 ["C-09","Facilities Technician",11,"HVAC, lifts, contractor supervision","HVAC Level 3, IOSH",8,"Kuwait City","Not screened"],
 ["C-10","Facilities Manager",8,"Contractor management, budgets, compliance","IOSH, NEBOSH",12,"Salmiya","Not screened"],
 ["C-11","Facilities Manager",5,"Maintenance planning, reporting","IOSH",6,"Kuwait City","Not screened"],
 ["C-12","Facilities Manager",14,"Multi-site operations, capital projects","NEBOSH, PMP",12,"Kuwait City","Not screened"],
]
sheet("Candidates",
      ["Candidate ID","Role Applied","Years Experience","Key Skills","Certifications",
       "Notice Period (weeks)","Location","Screening Status"],
      CAND, [14,22,17,40,26,21,15,16],
      title="Candidate screening sheet",
      note="Sample data — fictional candidates. No real applicants. IDs only, deliberately no names.")

# ─────────────────────────────── README (first sheet)
wb.remove(wb["Sheet"])
ws = wb.create_sheet("README", 0)
lines = [
 ("Alshaya & Al-Injaz — AI in Workflow for Professionals", TITLE_FONT),
 ("Day 2 lab dataset", Font(name="Arial", size=11, bold=True)),
 ("", BODY_FONT),
 ("SAMPLE DATA — TRAINING USE ONLY.", Font(name="Arial", size=10, bold=True, color="8B1E2A")),
 ("Every property, ticket, budget line, survey response and candidate in this workbook is invented", NOTE_FONT),
 ("for the workshop. It is not an Alshaya or Al-Injaz record. Safe to upload to Claude.", NOTE_FONT),
 ("", BODY_FONT),
 ("Sheets in this workbook", Font(name="Arial", size=11, bold=True)),
 ("Occupancy & Income   — 18 rows. Monthly occupancy, rent and collections, Jan–Jun 2026.", BODY_FONT),
 ("Maintenance Log      — 68 rows. Every ticket raised Apr–Jun 2026, open and closed.", BODY_FONT),
 ("Budget vs Actual     — 5 rows. Q3 operating budget, planned against actual.", BODY_FONT),
 ("Employee Survey      — 12 rows. Satisfaction scores across four departments.", BODY_FONT),
 ("Candidates           — 12 rows. Anonymous candidate screening sheet for the HR exercise.", BODY_FONT),
 ("", BODY_FONT),
 ("Deliberately no formulas", Font(name="Arial", size=11, bold=True)),
 ("Percentages and day counts are given; every total, average and variance is left for you", BODY_FONT),
 ("to calculate. That calculation is the exercise.", BODY_FONT),
 ("", BODY_FONT),
 ("This workbook is consistent with the Day 1 Quarterly Property Operations Report.", NOTE_FONT),
]
for i, (txt, f) in enumerate(lines, 1):
    c = ws.cell(i, 1, txt); c.font = f
ws.column_dimensions["A"].width = 100

wb.save(OUT)
print("wrote", OUT)
